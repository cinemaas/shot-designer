#!/usr/bin/env python3
"""Local server for the Shot Designer clone.

Serves the app and exposes a small file API over the real
"~/Documents/Shot Designer Scenes" folder, so the clone opens the same .hcw
scenes the original app uses. Writes always leave a timestamped backup behind.
"""
import http.server, socketserver, json, os, re, shutil, time, urllib.parse, posixpath
import base64, tempfile

PORT = 8769
APP_DIR = os.path.dirname(os.path.abspath(__file__))
SCENES = os.path.expanduser("~/Documents/Shot Designer Scenes")
BACKUPS = os.path.join(SCENES, ".sdclone-backups")
DATA = os.path.join(SCENES, ".sdclone")          # our own library, beside the scenes


def safe(rel):
    """Resolve a client-supplied path inside SCENES, or raise."""
    rel = posixpath.normpath("/" + (rel or "").lstrip("/")).lstrip("/")
    full = os.path.realpath(os.path.join(SCENES, rel))
    if not full.startswith(os.path.realpath(SCENES)):
        raise ValueError("path escapes scenes folder")
    return full


def data_path(key):
    """A JSON blob under .sdclone/, addressed by a safe slug path."""
    if not re.fullmatch(r"[A-Za-z0-9 _./-]{1,200}", key or ""):
        raise ValueError("bad key")
    full = os.path.realpath(os.path.join(DATA, key + ".json"))
    if not full.startswith(os.path.realpath(DATA)):
        raise ValueError("path escapes data folder")
    return full


def listing(rel):
    full = safe(rel)
    folders, scenes = [], []
    for name in sorted(os.listdir(full), key=str.lower):
        if name.startswith("."):
            continue
        p = os.path.join(full, name)
        if os.path.isdir(p):
            folders.append(name)
        elif name.lower().endswith(".hcw"):
            st = os.stat(p)
            scenes.append({"name": name, "size": st.st_size, "mtime": st.st_mtime})
    return {"path": rel, "folders": folders, "scenes": scenes}


DRIVE = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-cinemaasinc@gmail.com/My Drive")
SHEET_DIR = os.path.join(DRIVE, "Shot Lists")


def drive_dest(rel):
    """A destination inside Drive, or the default folder."""
    rel = (rel or "Shot Lists").strip().strip("/")
    full = os.path.realpath(os.path.join(DRIVE, rel))
    if not full.startswith(os.path.realpath(DRIVE)):
        raise ValueError("destination must be inside Google Drive")
    return full


def tab_name(name):
    """Sheet names can't carry the characters a scene path can."""
    clean = re.sub(r"[\\/*?:\[\]]", "-", name).strip() or "Scene"
    return clean[:31]


def write_shot_sheet(data):
    """A formatted shot list with a frame per row, dropped straight into Drive.

    Google Drive syncs the folder, and Sheets opens .xlsx directly, so it's
    there ready to share with departments without an upload step. One tab per
    scene, plus an "All Shots" tab a department can filter in one go.
    """
    from openpyxl import Workbook

    sheets = data.get("sheets")
    if not sheets:
        sheets = [{"name": data.get("scene") or "Scene", "shots": data.get("shots") or []}]

    title = re.sub(r"[^\w .-]", "", data.get("title") or sheets[0]["name"] or "Shot List")
    dest = drive_dest(data.get("dest"))

    wb = Workbook()
    wb.remove(wb.active)
    tmp = []

    if len(sheets) > 1:
        flat = []
        for sh in sheets:
            for shot in sh["shots"]:
                flat.append({**shot, "scene": sh["name"]})
        fill_sheet(wb.create_sheet(tab_name("All Shots")), "All Shots", flat, tmp,
                   with_scene=True)

    used = set()
    for sh in sheets:
        name = tab_name(sh["name"])
        n, base = 2, name
        while name in used:
            name = tab_name(f"{base} {n}"); n += 1
        used.add(name)
        fill_sheet(wb.create_sheet(name), sh["name"], sh["shots"], tmp)

    os.makedirs(dest, exist_ok=True)
    path = os.path.join(dest, f"{title} shot list.xlsx")
    wb.save(path)
    for f in tmp:
        try: os.remove(f)
        except OSError: pass
    return {"ok": True, "path": path, "folder": os.path.relpath(dest, DRIVE),
            "scenes": len(sheets),
            "shots": sum(len(sh["shots"]) for sh in sheets)}


def fill_sheet(ws, scene, shots, tmp, with_scene=False):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    headers = (["#", "Scene"] if with_scene else ["#"]) + \
              ["Cam", "Shot", "Type", "Lens", "Notes", "Overhead"]
    widths = ([5, 16] if with_scene else [5]) + [8, 34, 9, 8, 40, 46]
    navy = PatternFill("solid", fgColor="255681")
    thin = Side(style="thin", color="D6DBE0")
    edge = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = scene
    ws["A1"].font = Font(size=15, bold=True, color="255681")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 26

    for i, (head, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=2, column=i, value=head)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = navy
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    notes_col = len(headers) - 1
    img_col = len(headers)
    for n, shot in enumerate(shots, start=1):
        row = n + 2
        values = ([n, shot.get("scene", "")] if with_scene else [n]) + [
            shot.get("camera", ""), shot.get("shot", ""),
            shot.get("type", ""), shot.get("lens", ""), shot.get("notes", "")]
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=(i == notes_col))
            c.border = edge
        ws.cell(row=row, column=img_col).border = edge
        ws.row_dimensions[row].height = 132

        png = shot.get("png") or ""
        if png.startswith("data:image"):
            raw = base64.b64decode(png.split(",", 1)[1])
            f = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            f.write(raw); f.close(); tmp.append(f.name)
            img = XLImage(f.name)
            scale = min(320 / img.width, 172 / img.height, 1)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            ws.add_image(img, f"{get_column_letter(img_col)}{row}")



class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *a, **kw):
        super().__init__(*a, directory=APP_DIR, **kw)

    def log_message(self, fmt, *args):
        pass

    def send_json(self, obj, code=200):
        body = json.dumps(obj).encode()
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        u = urllib.parse.urlparse(self.path)
        q = urllib.parse.parse_qs(u.query)
        try:
            if u.path == "/api/list":
                return self.send_json(listing(q.get("path", [""])[0]))
            if u.path == "/api/scene":
                full = safe(q.get("path", [""])[0])
                with open(full, encoding="utf-8", errors="replace") as f:
                    return self.send_json({"path": q["path"][0], "xml": f.read()})
            if u.path == "/api/data":
                full = data_path(q.get("key", [""])[0])
                if not os.path.exists(full):
                    return self.send_json({"value": None})
                with open(full, encoding="utf-8") as f:
                    return self.send_json({"value": json.load(f)})
            if u.path == "/api/data-list":
                folder = os.path.join(DATA, q.get("folder", [""])[0])
                names = []
                if os.path.isdir(folder):
                    names = sorted((n[:-5] for n in os.listdir(folder)
                                    if n.endswith(".json")), key=str.lower)
                return self.send_json({"names": names})
        except Exception as e:
            return self.send_json({"error": str(e)}, 400)
        # Never let the browser cache the app while we're iterating on it.
        self.send_header_nocache = True
        return super().do_GET()

    def end_headers(self):
        if getattr(self, "send_header_nocache", False):
            self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_POST(self):
        u = urllib.parse.urlparse(self.path)
        n = int(self.headers.get("Content-Length", 0))
        try:
            data = json.loads(self.rfile.read(n) or b"{}")
            if u.path == "/api/save":
                full = safe(data["path"])
                os.makedirs(os.path.dirname(full), exist_ok=True)
                if os.path.exists(full):
                    os.makedirs(BACKUPS, exist_ok=True)
                    stamp = time.strftime("%Y%m%d-%H%M%S")
                    flat = data["path"].replace("/", "_")
                    shutil.copy2(full, os.path.join(BACKUPS, f"{stamp}_{flat}"))
                with open(full, "w", encoding="utf-8") as f:
                    f.write(data["xml"])
                return self.send_json({"ok": True, "path": data["path"]})
            if u.path == "/api/data":
                full = data_path(data["key"])
                os.makedirs(os.path.dirname(full), exist_ok=True)
                with open(full, "w", encoding="utf-8") as f:
                    json.dump(data["value"], f, indent=1)
                return self.send_json({"ok": True})
            if u.path == "/api/data-delete":
                full = data_path(data["key"])
                if os.path.exists(full):
                    os.remove(full)
                return self.send_json({"ok": True})
            if u.path == "/api/copy":
                src, dst = safe(data["from"]), safe(data["to"])
                if os.path.exists(dst):
                    return self.send_json({"error": "a scene with that name already exists"}, 400)
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                shutil.copy2(src, dst)
                return self.send_json({"ok": True, "path": data["to"]})
            if u.path == "/api/shotsheet":
                return self.send_json(write_shot_sheet(data))
            if u.path == "/api/mkdir":
                os.makedirs(safe(data["path"]), exist_ok=True)
                return self.send_json({"ok": True})
        except Exception as e:
            return self.send_json({"error": str(e)}, 400)
        return self.send_json({"error": "unknown endpoint"}, 404)


class Server(socketserver.ThreadingTCPServer):
    allow_reuse_address = True
    daemon_threads = True


if __name__ == "__main__":
    print(f"Shot Designer clone  ->  http://localhost:{PORT}")
    print(f"Scenes folder: {SCENES}")
    Server(("127.0.0.1", PORT), Handler).serve_forever()
